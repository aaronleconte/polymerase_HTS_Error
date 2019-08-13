"""Calculating Error Rate from FASTQ files"""
import gzip
import os
import shutil
import re
from Bio import SeqIO
from Bio import Align
from Bio import pairwise2
from Bio.pairwise2 import format_alignment
from pandas import DataFrame

## _____________________________________________________________________________________________________________________________________ ##
# **USER INPUT**
# 1. Start by entering the file name and other basic info.  It is REALLY important to remember to do this each time!

# Put in the experiment name (e.g. MS-098-1) here:
experiment = "AV2-133-1"
# Add a description of1 the template name (e.g. K021)
templatename = "K021"
# Add a description of the experiment (e.g. testing P1 (GTCGTGAT) and template control (ACCACTGT) to reproduce past work)
description = "uses 0718 code, tests conditions for Q5 RT/amp (SF, ?? cycles) on P1 with all F (GTCGTGAT), P1 with AzA (ACCACTGT), 4-6 with all F (TGGATCTG), 4-6 with AzA (CCGTTTGT), template (TGCTGGGT)"

#2. Select the construct that you are using here.  Whichever one that you DO NOT use should be marked by hashtags (this makes the program not read the file

#If using K021 template and 359/360, use the info below (if not, hashtag it out):
primer1 = "TTCTCGCCAAAGACCTGAGCGTTCTGGCCCTGAGGATGCTGTCTACACGCA" #forward primer sequence
primer2 = "TTTCCCGCAAGGAGCCCATGTGGGCCGATCTTCTGAGTGACGATTCAAGGCT" #reverse primer sequence
barcodeSeq = ["GTCGTGAT", "ACCACTGT", "TGGATCTG", "CCGTTTGT", "TGCTGGGT", "GAGGGGTT", "AGGTTGGG", "GTGTGGTG", "TGGGTTTC"] # list of barcodes to look for -continue adding to list for more barcodes
template = "AGCTTACATTAAGACTCGCCATGTTACGATCTGCCTCAGGTAAGTAC" #template strand

#If using T57 construct, use the info below (if not, hashtag it out):
#primer1 = "TCATCAAAGACCTGAGCGTTCTGGCACTGAGGATGATGTCAAGATGCAAGTGACGATTCAAGGCT" #forward primer sequence
#primer2 = "CAGTCAAGGACCTCATGTGGGCCGTTCTTCTGAGTGACGATTCAAGGTATGCTGTCTACACGC" #reverse primer sequence
#barcodeSeq = ["GTCGTGAT", "ACCACTGT", "TGGATCTG", "CCGTTTGT", "TGCTGGGT", "GAGGGGTT", "AGGTTGGG", "GTGTGGTG", "TGGGTTTC"] # list of barcodes to look for -continue adding to list for more barcodes
#template = "TACCTGAGGCAGATCGTAACATGCT" #template strand

# # _____________________________________________________________________________________________________________________________________ ##


# Unzip the .gz files in your current directory
def unzipGZ():
    """ This function unzips all .gz files IN YOUR CURRENT DIRECTORY 
    and creates a list of unzipped filenames.
    parameters: none
    returns: returns a directory of uzipped files"""
    zippedFiles = []
    unzippedFiles = []

    dirFiles = os.listdir('.') # directory of all files in current folder
    for files in dirFiles: 
        if '.gz' in files: # filter so we only have the zipped files (.gz files)
            zippedFiles.append(files)

    # Unzip the files and make a list of unzipped files
    if len(zippedFiles) > 1:
        for filename in zippedFiles:
            # Unzip
            with gzip.open(filename, 'rb') as f_in:
                with open(filename[0:-3], 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            #Append to list of unzipped fies
            unzippedFiles.append(filename[0:-3])
    #condition if we are only unzipping one file
    else:
        with gzip.open(zippedFiles[0], 'rb') as f_in:
            with open(zippedFiles[0][0:-3], 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)
        unzippedFiles.append(zippedFiles[0][0:-3]) 
    return unzippedFiles


def trim_barcodes(records, barcode):
    """Trims barcode sequences.

    This is a generator function, the records argument should
    be a list or iterator returning SeqRecord objects.
    """
    barcodeRecords = []
    noBarcode = []
    len_barcode = len(barcode) #cache this for later
    for record in records:
        if record[1] != -1:
            index = record[0].find(barcode)
            if index != -1:
                record = [record[0][0:index]+record[0][index+len_barcode:], record[1]]
                barcodeRecords.append(record)
            else:
                index = record[0].find(reverse_complement(barcode))
                if index != -1:
                    record = [record[0][0:index]+record[0][index+len_barcode:], record[1]]
                    barcodeRecords.append(record)
                else:
                    noBarcode.append(record)
    allRecords = [noBarcode, barcodeRecords]
    return allRecords


def parseBarcodes(allReads, barcodeSeq, final_reads):
    """ This function parses through the specified files and finds instances of each barcode.
    It then creates a fastq file for the sequences with barcodes and trims out the barcode sequence by calling trim_barcodes.
    Files are saved in the directory with the filename + name of the barcode:
        i.e. "AL01-MS098-1_s1_L001_r1_001_ACCACTGT.fastq" 
    This function also calls trimPrimers to trim off the primers of the barcode sequence

    parameter: barcodeSeq - the barcodes to search for
    parameter: primer1 - the forward to the template strand
    parameter: primer2 - the reverse to the template strand
    parameter: filename - the name of the fastqfile that has barcoded sequences
    return: none"""
    for barcode in barcodeSeq:
        barcodeName = barcode
        trimmed_reads = trim_barcodes(final_reads, barcode)
        final_reads = trimmed_reads[0]
        allReads.update({"No Barcode" : final_reads})
        allReads.update({barcodeName : trimmed_reads[1]})

  


def parseDirection(frontPrimer, endPrimer, filename, finalFilename):
    """ This function parses through the specified file and seperates out the forward and reverse reads.

    parameter: primer - the forward or reverse to the template strand
    parameter: filename - the name of the fastq file to seperate out
    parameter: finalFilename - the name of the final, parsed file
    return: none"""
    primer_reads = (rec for rec in \
                    SeqIO.parse(filename, "fastq") \
                    if frontPrimer[-5:] in rec and reverse_complement(endPrimer)[0:5] in rec) #and rec.seq.endswith(reverse_complement(primer1))) #'CCCAACCt'in rec)
    count = SeqIO.write(primer_reads, finalFilename, "fastq")
    print("Saved %i reads" % count)


def trimPrimers2(frontPosition, endPosition, filename):
    """ This function trims the primers

    parameter: frontPosition - the end of the primer sequence in the front
    parameter: endPosition - the start of the primemr squence in the end
    parameter: filename - the name of the parsed file
    return: an array of trimmed sequences"""
    sequence_reads = []
    index = 0
    for rec in SeqIO.parse(filename, "fastq"):
        sequence = rec.seq
        ided_sequence = [rec.id, sequence[frontPosition[index][2]:endPosition[index][1]]]
        sequence_reads.append(ided_sequence)
        index=index+1
    return sequence_reads


def search_fastq(pattern, filename):
    """ returns the position of the primer snippet nearest the sequence

    parameter: pattern - the portion of the primer sequence
    parameter: filename - the name of the parsed file
    return: an array of indices where the sequence starts/ends"""
    start_end_pos = []
    for record in SeqIO.parse(filename, "fastq"):
            chrom = record.id
            count = 0
            for match in re.finditer(pattern, str(record.seq)):
                if count == 0:
                    start_pos = match.start()
                    end_pos = match.end()
                    pair = [chrom, start_pos, end_pos]
                    start_end_pos.append(pair)
                count = count+1
    return start_end_pos

        
def fileList(content):
    """ Creates and returns a list of files based on string content"""
    listFiles = []
    dirFiles = os.listdir('.') # directory of all files in current folder
    for files in dirFiles: 
        if content in files: # filter so we only have the zipped files (.gz files)
            listFiles.append(files)
    return listFiles

#pairwise -- use this to find the deletions, insertions and misincorps 
def pairwiseComp(seq, synthStrand):
    """ This function takes in a sequence and compares it to the reference strand
    using a pairwise comparison from biopython. It scores the comparison based on correct pairs,
    incorrect pairs, gap openning, and gap extension to determine the best match 
    between sequence and reference strand.

    parameter: seq - sequence compared to refernce strand
    parameter: synthStrand - reference strand
    return: a multi-dimensional array with alignments and scores"""
    alignments = None
    for alignments in pairwise2.align.globalms(seq, synthStrand, 5, -4, -3, -.1): #5 for every correct pair, -4 for an incorrect pair, -3 for opening a gap, -.1 when extending it
        format_alignment(*alignments)
    return alignments 


def mistakes(alignments):
    """ This function finds the mistakes in the sequence by determining where the alignments from
    pairwiseComp do not align (denoted by a - in the refernce strand or sequence)

    parameter: alignments - the output of pairwiseComp
    return: an array of matches/mismatches , where an insertion is denoted by i, deletion by d and misincorp by m"""
    match = []
    if alignments[0] != alignments[1]:
        for a, b in zip(alignments[0],alignments[1]): # a is the sequence in question and b is the refernce
            if a == b:
                    match.append('|')
            elif a == '-': 
                    match.append('d')
            elif b == '-':
                    match.append('i')
            else:
                    match.append('m')
        return match

    """Below is code that prints out the sequence aligned with the refernce sequence that visually
    shows mistakes -- pretty cool to look at but not very necessary except for troubleshooting
    match = []

    for a, b in zip(alignments[0],alignments[1]):
            if a == b:
                    match.append('|')
            elif a == '-' or b == '-':
                    match.append(' ')
            else:
                    match.append('x')

    m="".join(match)
    s=[]
    s.append(alignments[0]+'\n')
    s.append(m+'\n')
    s.append(alignments[1])

    alignedSeqs="".join(s)
    print('\n')
    print(alignedSeqs)
    
    return alignments"""

def aggregateErrorRate(comparedSequence, templateLength):
    """ This function computes the aggregate error rate for all and each type of mistake

    parameter: comparedSequence - the output of mistakes
    parameter: templateLength - the length of the template strand
    return: an array of errors for all the sequences which is presented in
    the following order: [total bases, total number of errors, total error rate, total number of insertions, insertion error rate,
    total number of deletions, deletion error rate, total number of misincorperations, misincorp erorr rate, nonmatchReads, matchReads]"""
    insCount = delCount = misCount = totalBases = totalErrorCount = matchReads = nonmatchReads = 0
    for seq in comparedSequence:
        if seq != None: #None means the two sequences matched -- no need to check those
            nonmatchReads+=1
            for pos in seq:
                if pos == 'i':
                    totalErrorCount=totalErrorCount+1
                    insCount=insCount+1
                elif pos == 'd':
                    totalErrorCount=totalErrorCount+1
                    delCount=delCount+1
                elif pos == 'm':
                    totalErrorCount=totalErrorCount+1
                    misCount=misCount+1
                totalBases+=1
        else: #still add bases matching sequences to base count
            matchReads+=1
            totalBases+=templateLength
    return [totalBases, totalErrorCount, totalErrorCount/(totalBases*1.0), insCount, insCount/(totalBases*1.0), delCount, delCount/(totalBases*1.0), misCount, misCount/(totalBases*1.0), nonmatchReads, matchReads]

def positionalErrorRate(comparedSequence, position, trimmedAlignment):
    """ This function computes the positional error rate for all and each type of mistake

    parameter: comparedSequence - the output of mistakes
    parameter: position - the index of the position in the sequence
    return: an array  of errors for a specific position, presented in
    the following order: [total bases, total number of errors, total error rate, total number of insertions, insertion error rate,
    total number of deletions, deletion error rate, total number of misincorperations, misincorp erorr rate, ID of base at that position]"""
    insCount = delCount = misCount = totalErrorCount = totalBases = index = 0
    baseID = []
    for seq in comparedSequence:
        if seq != None:
            if seq[position] == 'i':
                totalErrorCount=totalErrorCount+1
                insCount=insCount+1
            elif seq[position] == 'd':
                totalErrorCount=totalErrorCount+1
                delCount=delCount+1
            elif seq[position] == 'm':
                totalErrorCount=totalErrorCount+1
                misCount=misCount+1
                baseID.append(trimmedAlignment[index][0][position])
        totalBases+=1
        index+=1
    return [totalBases, totalErrorCount, totalErrorCount/(totalBases*1.0), insCount, insCount/(totalBases*1.0), delCount, delCount/(totalBases*1.0), misCount, misCount/(totalBases*1.0), baseID]

def aggregateErrorSpectrum(positionalErrorRates, synthStrand, baseRef):
    """ This function computes the aggregate error spectrum for each type of mistake

    parameter: positionalErrorRates - the output of positionalErrorRates
    parameter: synthStrand - the reverse complement of the template strand
    paramater: baseRef - A, T, C, G -- the base we are finding the aggregate error spectrum for
    return: an array  of errors for the base specifed, presented in
    the following order: [total bases, total number of errors, total error rate, total number of insertions, insertion error rate,
    total number of deletions, deletion error rate, total number of misincorperations, misincorp erorr rate]"""
    index = insCount = delCount = misCount = totalErrorCount = totalBases = totalA = totalT = totalC = totalG = 0
    
    for pos in positionalErrorRates:
        if synthStrand[index] == baseRef:
            insCount += pos[3]
            delCount += pos[5]
            misCount += pos[7]
            totalErrorCount += pos[1]
            totalBases += pos[0]
            if misCount != 0:
                for base in pos[9]:
                    if base == 'A':
                        totalA+=1
                    elif base == 'T':
                        totalT+=1
                    elif base == 'C':
                        totalC+=1
                    elif base == 'G':
                        totalG+=1
        index+=1
    return [totalBases, totalErrorCount, totalErrorCount/(totalBases*1.0), insCount, insCount/(totalBases*1.0), delCount, delCount/(totalBases*1.0), misCount, misCount/(totalBases*1.0), totalA, totalT, totalC, totalG]


def reverse_complement(dna):
    """ find the reverse_complement of the sequence """
    complement = {'A': 'T', 'C': 'G', 'G': 'C', 'T': 'A'}
    return ''.join([complement[base] for base in dna[::-1]])

""" old trim Primers from 190508
def trimPrimers(primer1, primer2, alignments):
    #find primer sequence of template strand for primer 1 and 2
    trimmedStrands = []
    forwardCount = 0
    reverseCount = 0
    for strand in alignments:
        while primer1 is not "":
            if strand[1][forwardCount] == "-":
                forwardCount+=1
            else:
                forwardCount+=1
                primer1 = primer1[1:]
        while primer2 is not "":
            if strand[1][0-reverseCount+1] == "-":
                reverseCount+=1
            else:
                reverseCount+=1
                primer2 = primer2[1:]
        trimmedStrands.append((strand[0][forwardCount-1:len(strand[0])-reverseCount], strand[1][forwardCount-1:len(strand[0])-reverseCount]))
    return trimmedStrands
    """
        

def output(key, allMistakes, template, trimmedAlignment):
    """ displays the output of the aggregate and spectrum error functions for testing --
    this will chnage to export an excel document """

    agEr= aggregateErrorRate(allMistakes, len(template))
    """agOut = {'Overview': ['All Errors:', 'Insertions:', 'Deletions:', 'Misincorperations:'],
        'Number of Errors': [agEr[1],agEr[3],agEr[5],agEr[7]],
        'Error Rate (per 1000 base pairs)': [agEr[2]*1000, agEr[4]*1000, agEr[6]*1000, agEr[8]*1000]
        }
    
    df = DataFrame(agOut, columns= ['Overview', 'Number of Errors', 'Error Rate (per 1000 base pairs)'])
    export_excel = df.to_excel (r'export_dataframe.xlsx', index = None, header=True) #Don't forget to add '.xlsx' at the end of the path
    """
    print("Unmatched Reads: %i\n" %agEr[9])
    print("Matched Reads: %i\n" %agEr[10])

    print("writing the file!")


    #Excel: write Excel file 
    import openpyxl
    wb=openpyxl.Workbook()

    #Excel: create Aggregate column headings and row headings
    wb.create_sheet("Aggregate",0)
    wb["Aggregate"].cell(1,1).value=experiment
    wb["Aggregate"].cell(2,1).value=templatename
    wb["Aggregate"].cell(3,1).value=description
    wb["Aggregate"].cell(4,1).value="Unmatched Reads:"+str(agEr[9])
    wb["Aggregate"].cell(5,1).value="Matched Reads:"+str(agEr[10])


    wb["Aggregate"].cell(12,1).value="Reads"
    wb["Aggregate"].cell(13,1).value="Nucleotides"
    wb["Aggregate"].cell(14,1).value="Total Errors"
    wb["Aggregate"].cell(15,1).value="Insertions"
    wb["Aggregate"].cell(16,1).value="Deletions"
    wb["Aggregate"].cell(17,1).value="Misincorporations"
    wb["Aggregate"].cell(11,2).value="Total"
    wb["Aggregate"].cell(11,3).value="Rate"
    wb["Aggregate"].cell(11,4).value="Per thousand base pairs"

    #Excel: Write the Excel file
    #Still need to write in some of the data that were not part of original file output
    wb["Aggregate"].cell(14,2).value=agEr[1]
    wb["Aggregate"].cell(14,3).value=agEr[2]
    wb["Aggregate"].cell(14,4).value=1000*agEr[2]
    wb["Aggregate"].cell(15,2).value=agEr[3]
    wb["Aggregate"].cell(15,3).value=agEr[4]
    wb["Aggregate"].cell(15,4).value=1000*agEr[4]
    wb["Aggregate"].cell(16,2).value=agEr[5]
    wb["Aggregate"].cell(16,3).value=agEr[6]
    wb["Aggregate"].cell(16,4).value=1000*agEr[6]
    wb["Aggregate"].cell(17,2).value=agEr[7]
    wb["Aggregate"].cell(17,3).value=agEr[8]
    wb["Aggregate"].cell(17,4).value=1000*agEr[8]

    #Excel: write the position file (positions themselves get added during recording of data)
    wb.create_sheet("Position",1)
    wb["Position"].cell(1,2).value="Correct base"
    wb["Position"].cell(1,3).value="Errors ptb"
    wb["Position"].cell(1,4).value="Insertions ptb"
    wb["Position"].cell(1,5).value="Deletions ptb"
    wb["Position"].cell(1,6).value="Misincorporations ptb"
    wb["Position"].cell(1,8).value="Total Errors"
    wb["Position"].cell(1,9).value="Total Insertions"
    wb["Position"].cell(1,10).value="Total Deletions"
    wb["Position"].cell(1,11).value="Misincorporations"

    indexCount = 0
    posEr = []
    for base in template:
        posEr.append(positionalErrorRate(allMistakes, indexCount, trimmedAlignment))
        wb["Position"].cell(indexCount+2,1).value=indexCount+1
        wb["Position"].cell(indexCount+2,2).value=reverse_complement(template)[indexCount]
        wb["Position"].cell(indexCount+2,8).value=posEr[indexCount][1]
        wb["Position"].cell(indexCount+2,3).value=round(posEr[indexCount][2]*1000,2)
        wb["Position"].cell(indexCount+2,9).value=posEr[indexCount][3]
        wb["Position"].cell(indexCount+2,4).value=round(posEr[indexCount][4]*1000,2)
        wb["Position"].cell(indexCount+2,10).value=posEr[indexCount][5]
        wb["Position"].cell(indexCount+2,5).value=round(posEr[indexCount][6]*1000,2)
        wb["Position"].cell(indexCount+2,11).value=posEr[indexCount][7]
        wb["Position"].cell(indexCount+2,6).value=round(posEr[indexCount][8]*1000,2)
        indexCount+=1

    specEr = []
    #Excel to write sheet
    wb.create_sheet("Spectrum",2)
    wb["Spectrum"].cell(1,2).value="Errors ptb"
    wb["Spectrum"].cell(1,3).value="Insertions ptb"
    wb["Spectrum"].cell(1,4).value="Deletions ptb"
    wb["Spectrum"].cell(1,5).value="Misincorporations ptb"
    wb["Spectrum"].cell(1,7).value="Total Errors"
    wb["Spectrum"].cell(1,8).value="Total Insertions"
    wb["Spectrum"].cell(1,9).value="Total Deletions"
    wb["Spectrum"].cell(1,10).value="Misincorporations"

    wb.create_sheet("Mistakes",3)
    wb["Mistakes"].cell(1,1).value="Original Base"
    wb["Mistakes"].cell(1,2).value="Changed Base"
    wb["Mistakes"].cell(1,3).value="Reads"
    
    indexCount = 0
    for base in ['A', 'T', 'C', 'G']:
        specEr.append(aggregateErrorSpectrum(posEr, reverse_complement(template), base))
        wb["Spectrum"].cell(indexCount+2,1).value=base
        wb["Spectrum"].cell(indexCount+2,7).value=specEr[indexCount][1]
        wb["Spectrum"].cell(indexCount+2,2).value=round(specEr[indexCount][2]*1000,2)
        wb["Spectrum"].cell(indexCount+2,8).value=specEr[indexCount][3]
        wb["Spectrum"].cell(indexCount+2,3).value=round(specEr[indexCount][4]*1000,2)
        wb["Spectrum"].cell(indexCount+2,9).value=specEr[indexCount][5]
        wb["Spectrum"].cell(indexCount+2,4).value=round(specEr[indexCount][6]*1000,2)
        wb["Spectrum"].cell(indexCount+2,10).value=specEr[indexCount][7]
        wb["Spectrum"].cell(indexCount+2,5).value=round(specEr[indexCount][8]*1000,2)
        if base == 'A':
            wb["Mistakes"].cell(2,1).value=base
            wb["Mistakes"].cell(3,1).value=base
            wb["Mistakes"].cell(4,1).value=base
            wb["Mistakes"].cell(2,2).value='T'
            wb["Mistakes"].cell(3,2).value='C'
            wb["Mistakes"].cell(4,2).value='G'
            wb["Mistakes"].cell(2,3).value=specEr[indexCount][10]
            wb["Mistakes"].cell(3,3).value=specEr[indexCount][11]
            wb["Mistakes"].cell(4,3).value=specEr[indexCount][12]
        elif base == 'T':
            wb["Mistakes"].cell(5,1).value=base
            wb["Mistakes"].cell(6,1).value=base
            wb["Mistakes"].cell(7,1).value=base
            wb["Mistakes"].cell(5,2).value='A'
            wb["Mistakes"].cell(6,2).value='C'
            wb["Mistakes"].cell(7,2).value='G'
            wb["Mistakes"].cell(5,3).value=specEr[indexCount][9]
            wb["Mistakes"].cell(6,3).value=specEr[indexCount][11]
            wb["Mistakes"].cell(7,3).value=specEr[indexCount][12]
        elif base == 'C':
            wb["Mistakes"].cell(8,1).value=base
            wb["Mistakes"].cell(9,1).value=base
            wb["Mistakes"].cell(10,1).value=base
            wb["Mistakes"].cell(8,2).value='A'
            wb["Mistakes"].cell(9,2).value='T'
            wb["Mistakes"].cell(10,2).value='G'
            wb["Mistakes"].cell(8,3).value=specEr[indexCount][9]
            wb["Mistakes"].cell(9,3).value=specEr[indexCount][10]
            wb["Mistakes"].cell(10,3).value=specEr[indexCount][12]
        else:
            wb["Mistakes"].cell(11,1).value=base
            wb["Mistakes"].cell(12,1).value=base
            wb["Mistakes"].cell(13,1).value=base
            wb["Mistakes"].cell(11,2).value='A'
            wb["Mistakes"].cell(12,2).value='T'
            wb["Mistakes"].cell(13,2).value='C'
            wb["Mistakes"].cell(11,3).value=specEr[indexCount][9]
            wb["Mistakes"].cell(12,3).value=specEr[indexCount][10]
            wb["Mistakes"].cell(13,3).value=specEr[indexCount][11]
        indexCount+=1
    #need to add in what the mistakes are -- ie A->G"""
    
    #Excel: This saves the file
    wb.save("Output_0718_"+experiment+"_"+key+".xlsx")

def syncRecords(records, records2):
    """ Takes in two runs and returns the sequence of the first 
    record along with if the two match (will return -1 if not)"""
    print("Syncing in Process... This will take a few minutes")
    count2 = 0
    index = 0
    final_records = []
    for record in records:
        if 'N' not in records2[0].seq:
            alignment = [record.seq, record.seq.find(reverse_complement(records2[0].seq))]
            count2+=1
            final_records.append(alignment)
        records2 = records2[1:]
    print(count2)
    return final_records


def parseDirection2(final_records, frontPrimer, endPrimer):
    """Takes in the outcome of syncRecords and the two primers
    and returns the  reads """
    directional_reads = []
    count = 0
    for record in final_records:
        if record[1] != -1:
            if (record[0].find(frontPrimer[-5:]) != -1) and (record[0].find(reverse_complement(endPrimer)[0:5])): 
                count+=1
                read = [record[0], record[0].find(frontPrimer[-5:]), record[0].find(reverse_complement(endPrimer)[0:5])]
                directional_reads.append(read)
    print(count)
    return directional_reads


def trimPrimers3(directional_reads, template):
    """Takes in the outcome of parseDirection2
    and returns the  reads with trimmed primer """
    trimmed_reads = []
    for record in directional_reads:
        read = record[0][(record[1]+5):(record[2])]
        if len(read) < len(template)+(len(template)*0.15) and len(read) > len(template)-(len(template)*0.15):
            trimmed_reads.append(read)
    return trimmed_reads


def main():

    #unzip the files in your directory
    unzippedFiles = unzipGZ()

    print("Making lists of the two runs.")
    records = list(SeqIO.parse(unzippedFiles[0], "fastq"))
    records2 = list(SeqIO.parse(unzippedFiles[1], "fastq"))

    #if some experiments have barcodes, seperate them out based on barcode and then trim the 
    #primer sequence
    """if barcodeBool == 1:
    for filename in unzippedFiles:
        parseBarcodes(barcodeSeq, primer1, primer2, filename)"""

    final_records = syncRecords(records, records2)

    print("Parse out barcodes.")
    allReads = {}
    parseBarcodes(allReads, barcodeSeq, final_records)
    print("Parsing out forward and reverse reads and trimming primers for all reads.")

    forwardRead = template
    reverseRead = reverse_complement(template)

    for key in allReads:
        reads_forward = parseDirection2(allReads[key], primer1, primer2)
        reads_reverse = parseDirection2(allReads[key], primer2, primer1)
        if len(reads_forward) != 0:

            trimmed_reads_forward = trimPrimers3(reads_forward, template)
            trimmed_reads = trimPrimers3(reads_reverse, template)
            for read in trimmed_reads_forward:
                trimmed_reads.append(reverse_complement(read))

            alignment = []
            for seq in trimmed_reads:
                alignment.append(pairwiseComp(seq, reverseRead))
            
            allMistakes = []
            for strand in alignment:
                if strand != None:
                    allMistakes.append(mistakes(strand))
            
            output(key, allMistakes, template, alignment)
            
            """reads = [trimmed_reads_forward, trimmed_reads_reverse]
            allReads.update({key : reads})"""



if __name__ == "__main__":
    main()

"""with open('alignment.txt', 'w') as f:
    for item in alignment:
        f.write("%s\n" % item[0] + "%s\n\n" % item[1])"""